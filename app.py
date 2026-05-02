import smtplib
from email.mime.text import MIMEText
from flask import Flask, render_template, request, redirect, session, send_file, flash
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from io import BytesIO
import mysql.connector
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table as PdfTable, TableStyle as PdfTableStyle, Paragraph, Spacer

app = Flask(__name__)
app.secret_key = 'secret123'

# =========================
# CONEXIÓN BD
# =========================
def get_db_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="P_D_T_2026",
        database="moneyhome"
    )

def formatear_miles(valor):
    return f"{float(valor):,.0f}".replace(",", ".")

# =========================
# LOGIN
# =========================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT * FROM usuarios WHERE email=%s AND password=%s", (email, password))
        user = cursor.fetchone()

        if user:
            session['user_id'] = user['id']
            session['nombre'] = user['nombre']

            # 🔥 LÓGICA DE FLUJO
            cursor.execute("SELECT COUNT(*) AS total FROM banco WHERE user_id=%s", (user['id'],))
            resultado = cursor.fetchone()

            cursor.close()
            conn.close()

            if resultado['total'] == 0:
                # Usuario nuevo → no tiene banco
                return redirect('/banco')
            else:
                # Usuario ya configurado
                return redirect('/')

        else:
            flash("Credenciales incorrectas", "danger")

    return render_template('login.html')

#RECUPERAR CLAVE

import uuid

@app.route('/recuperar', methods=['GET', 'POST'])
def recuperar():
    if request.method == 'POST':
        email = request.form['email']

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT * FROM usuarios WHERE email=%s", (email,))
        user = cursor.fetchone()

        if user:
            token = str(uuid.uuid4())

            cursor.execute("""
            UPDATE usuarios
            SET reset_token=%s,
            reset_expira = DATE_ADD(NOW(), INTERVAL 1 HOUR)
            WHERE email=%s
        """, (token, email))
            conn.commit()

            link = f"http://127.0.0.1:5000/reset/{token}"

            # 🔥 ENVÍO REAL
            enviar_email(
                email,
                "Recuperación de contraseña - MoneyHome",
                f"Hola!\n\nHaz click aquí para recuperar tu contraseña:\n{link}"
            )

        cursor.close()
        conn.close()

        flash("Si el correo existe, se enviaron instrucciones.", "success")
        return redirect('/login')

    return render_template('recuperar.html')

# FUNCION PARA ENVIAR MAIL

def enviar_email(destinatario, asunto, cuerpo):
    remitente = "moneyhomemain@gmail.com"
    password = "xeef zefz tdjc fplj "

    mensaje = MIMEText(cuerpo)
    mensaje["Subject"] = asunto
    mensaje["From"] = remitente
    mensaje["To"] = destinatario

    try:
        servidor = smtplib.SMTP("smtp.gmail.com", 587)
        servidor.starttls()
        servidor.login(remitente, password)
        servidor.send_message(mensaje)
        servidor.quit()
    except Exception as e:
        print("❌ ERROR SMTP:")
        print(e)
        raise e
        

# =========================
# REGISTER
# =========================
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        email = request.form['email']

        # 🔍 VALIDAR SI YA EXISTE
        cursor.execute("SELECT id FROM usuarios WHERE email=%s", (email,))
        existente = cursor.fetchone()

        if existente:
            flash("Este correo ya está registrado", "danger")
            cursor.close()
            conn.close()
            return redirect('/register')

        # ✅ INSERTAR SI NO EXISTE
        cursor.execute("""
            INSERT INTO usuarios (nombre, email, password)
            VALUES (%s, %s, %s)
        """, (
            request.form['nombre'],
            email,
            request.form['password']
        ))

        conn.commit()
        cursor.close()
        conn.close()

        flash("Cuenta creada correctamente", "success")
        return redirect('/login')

    return render_template('register.html')

# =========================
# LOGOUT
# =========================
@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

# =========================
# DASHBOARD (CON FAMILIA + FLUJO PRO)
# =========================
@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect('/login')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # 🔥 VALIDAR SI TIENE BANCO
    cursor.execute(
        "SELECT id FROM banco WHERE user_id=%s LIMIT 1",
        (session['user_id'],)
    )
    banco = cursor.fetchone()

    # 👉 SI NO TIENE BANCO → FORZAR FLUJO
    if not banco:
        cursor.close()
        conn.close()
        return redirect('/banco')

    # =========================
    # INGRESOS Y GASTOS
    # =========================
    cursor.execute("""
        SELECT 
            COALESCE(SUM(CASE WHEN tipo_id = 1 THEN monto ELSE 0 END),0) AS ingresos,
            COALESCE(SUM(CASE WHEN tipo_id = 2 THEN monto ELSE 0 END),0) AS gastos
        FROM movimientos
        WHERE user_id = %s
    """, (session['user_id'],))

    data = cursor.fetchone()
    ingresos = data['ingresos']
    gastos = data['gastos']

    # =========================
    # 🔥 NUEVO: SALDO BANCOS
    # =========================
    cursor.execute("""
        SELECT COALESCE(SUM(monto),0) AS saldo_bancos
        FROM banco
        WHERE user_id = %s
    """, (session['user_id'],))

    saldo_bancos = cursor.fetchone()['saldo_bancos']

    # =========================
    # 🔥 SALDO FINAL REAL
    # =========================
    saldo = saldo_bancos + ingresos - gastos

    # =========================
    # MOVIMIENTOS
    # =========================
    cursor.execute("""
        SELECT m.*, c.nombre AS categoria, t.nombre AS tipo, u.nombre AS usuario
        FROM movimientos m
        LEFT JOIN categorias c ON m.categoria_id = c.id
        LEFT JOIN tipo_movimiento t ON m.tipo_id = t.id
        LEFT JOIN usuarios u ON m.user_id = u.id
        WHERE m.user_id = %s 
           OR u.familia_id = (
                SELECT familia_id FROM usuarios WHERE id = %s
           )
        ORDER BY m.fecha DESC
    """, (session['user_id'], session['user_id']))

    movimientos = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template('index.html',
        ingresos=ingresos,
        gastos=gastos,
        saldo=saldo,
        movimientos=movimientos
    )

# =========================
# MOVIMIENTOS (CON FAMILIA)
# =========================
@app.route('/mov')
@app.route('/mov/<tipo>', methods=['GET', 'POST'])
def mov(tipo=None):

    if 'user_id' not in session:
        return redirect('/login')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # CREAR
    if tipo == 'crear' and request.method == 'POST':
        tipo_map = {'ingreso':1,'gasto':2,'transferencia':3}

        try:
            cursor.execute("""
                INSERT INTO movimientos (user_id, monto, categoria_id, banco_id, tipo_id, fecha, descripcion)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
            """, (
                session['user_id'],
                request.form['monto'],
                request.form['categoria_id'],
                request.form['banco_id'],
                tipo_map.get(request.form['tipo_movimiento']),
                request.form['fecha'],
                request.form['descripcion']
            ))

            conn.commit()
            flash('Movimiento creado exitosamente', 'success')
            return redirect('/mov')
        except Exception as e:
            conn.rollback()
            flash(f'Error al crear movimiento: {str(e)}', 'danger')
            print(f"ERROR: {str(e)}")
        finally:
            cursor.close()
            conn.close()

    cursor.execute("SELECT * FROM categorias")
    categorias = cursor.fetchall()

    cursor.execute(
cursor.execute(
    """
    SELECT b.id,
           b.nombre_banco AS nombre,
           COALESCE(
               b.monto + SUM(
                   CASE
                       WHEN m.tipo_id = 1 THEN m.monto
                       WHEN m.tipo_id = 2 THEN -m.monto
                       ELSE 0
                   END
               ),
               b.monto
           ) AS saldo_actual
    FROM banco b
    LEFT JOIN movimientos m ON m.banco_id = b.id
    WHERE b.user_id = %s
    GROUP BY b.id, b.nombre_banco, b.monto
    ORDER BY b.id DESC
    """,
    (session['user_id'],)
)
    )
    bancos = cursor.fetchall()


    filtro = "AND m.tipo_id IN (1,2)"
    if tipo == 'ingresos':
        filtro = "AND m.tipo_id=1"
    elif tipo == 'gastos':
        filtro = "AND m.tipo_id=2"
    elif tipo == 'transferencias':
        filtro = "AND m.tipo_id=3"
    # mostrar solo los movimientos del usuario y su familia (si tiene)
    cursor.execute(f"""
        SELECT m.id, m.user_id, m.banco_id, m.monto, m.categoria_id, m.tipo_id, m.fecha, m.descripcion,
               c.nombre AS categoria,
               CASE 
                   WHEN m.tipo_id = 1 THEN 'ingreso'
                   WHEN m.tipo_id = 2 THEN 'gasto'
                   WHEN m.tipo_id = 3 THEN 'transferencia'
                   ELSE 'otro'
               END AS tipo,
               u.nombre AS usuario
        FROM movimientos m
        LEFT JOIN categorias c ON m.categoria_id = c.id
        LEFT JOIN usuarios u ON m.user_id = u.id
        WHERE m.user_id = %s
        {filtro}
        ORDER BY m.fecha DESC, m.id DESC
    """,(session['user_id'],))

    movimientos = cursor.fetchall()

    # Ingresos disponibles para poblar los selectores de transferencia.
    cursor.execute("""
        SELECT m.id, m.descripcion, m.monto, m.fecha, u.nombre AS usuario
        FROM movimientos m
        LEFT JOIN usuarios u ON m.user_id = u.id
        WHERE m.tipo_id = 1
          AND (m.user_id=%s 
           OR u.familia_id = (
                SELECT familia_id FROM usuarios WHERE id = %s
           ))
        ORDER BY m.fecha DESC, m.id DESC
    """, (session['user_id'], session['user_id']))
    ingresos_transferencia = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        'mov.html',
        movimientos=movimientos,
        categorias=categorias,
        bancos=bancos,
        tipo=tipo,
        ingresos_transferencia=ingresos_transferencia
    )

# =========================
# EDITAR MOVIMIENTO
# =========================
#Trae los datos de la tabla MOVIMIENTOS y CATEGORIAS para mostrar en la opcion editar.
@app.route('/mov/editar/<int:id>', methods=['GET','POST'])
def editar_movimiento(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Traer bancos para mostrar en el selector del formulario.
    cursor.execute("SELECT * FROM banco WHERE user_id=%s", (session['user_id'],))
    bancos = cursor.fetchall()

    #Obtener los bancos para mostrar en el selector del formulario.
    cursor.execute("""
    SELECT b.id, b.nombre_banco AS nombre
    FROM banco b
    WHERE b.user_id = %s
    ORDER BY b.id DESC
""", (session['user_id'],))

    # Validar que el movimiento a editar pertenece al usuario logueado.
    if request.method == 'POST':
        cursor.execute("""
            UPDATE movimientos
            SET descripcion=%s, monto=%s, categoria_id=%s, fecha=%s
            WHERE id=%s AND user_id=%s
        """, (
            request.form['descripcion'],
            request.form['monto'],
            request.form['categoria_id'],
            request.form['fecha'],
            id,
            session['user_id']
        ))

        conn.commit()
        return redirect('/mov')

    cursor.execute("SELECT * FROM movimientos WHERE id=%s",(id,))
    movimiento = cursor.fetchone()

    cursor.execute("SELECT * FROM categorias")
    categorias = cursor.fetchall()

    return render_template('editar_movimiento.html',
        movimiento=movimiento,
        categorias=categorias,
        bancos=bancos
    )
# ==================================
# BOTON EDITAR - ELIMNAR MOVIMIENTO
# ==================================

# Endpoint para mostrar la confirmación de eliminación de un movimiento.
@app.route('/mov/eliminar/<int:id>', methods=['GET'])
def confirmar_eliminar_movimiento(id):
    if 'user_id' not in session:
        return redirect('/login')
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM movimientos WHERE id=%s",(id,))
    movimiento = cursor.fetchone()

    return render_template('confirmar_eliminar_movimiento.html', movimiento=movimiento)

# Endpoint para eliminar el movimiento después de la confirmación.
@app.route('/mov/eliminar', methods=['POST'])
def eliminar_movimiento():
    if 'user_id' not in session:
        return redirect('/login')
    
    id = request.form['id']
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # 🔥 borrar pagos primero
        cursor.execute("DELETE FROM historial_pagos WHERE id_movimiento=%s", (id,))
        
        # 🔥 luego movimiento
        cursor.execute("DELETE FROM movimientos WHERE id=%s AND user_id=%s",(id,session['user_id']))
        
        conn.commit()
        flash("Movimiento eliminado correctamente", "success")

    except Exception as e:
        conn.rollback()
        flash("No se pudo eliminar el movimiento", "danger")
        print(e)

    cursor.close()
    conn.close()

    return redirect('/mov')

# =========================
#           PAGO
# =========================

# Endpoint para mostrar la página de pago, con el listado de gastos y bancos disponibles.
@app.route('/mov/pago')
def pago():
    if 'user_id' not in session:
        return redirect('/login')
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    # Traer bancos con saldo real
    cursor.execute("SELECT * FROM banco WHERE user_id=%s", (session['user_id'],))
    bancos = cursor.fetchall()

    # Traer gastos del usuario para mostrar en la página de pago.
    cursor.execute("""
        SELECT 
            m.id, 
            m.descripcion, 
            m.monto,
            m.monto_pagado,
            m.saldo_pendiente,
            c.nombre AS categoria
        FROM movimientos m
        LEFT JOIN categorias c ON m.categoria_id = c.id
        WHERE m.user_id = %s AND m.tipo_id = 2
        ORDER BY m.fecha DESC, m.id DESC
    """, (session['user_id'],))
    gastos = cursor.fetchall()

    return render_template('pago.html', bancos=bancos, gastos=gastos, movimiento=None, saldo_banco=None)
# ENDPOINT PARA REALIZAR EL PAGO DE UN GASTO. 
# Este endpoint maneja tanto la visualización de la página de pago para un gasto específico (GET) 
# como el procesamiento del pago cuando se envía el formulario (POST).
@app.route('/mov/pago/<int:id>', methods=['GET', 'POST'])
def pago_movimiento(id):
    if 'user_id' not in session:
        return redirect('/login')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Traer bancos con saldo real para mostrar en el selector del formulario de pago.
    cursor.execute("""
    SELECT 
        b.id, 
        b.nombre_banco AS nombre, 
        b.monto,
        COALESCE(
            b.monto + SUM(
                CASE
                    WHEN m.tipo_id = 1 THEN m.monto
                    WHEN m.tipo_id = 2 THEN -m.monto
                    ELSE 0
                END
            ),
            b.monto
        ) AS saldo_real
    FROM banco b
    LEFT JOIN movimientos m ON m.banco_id = b.id
    WHERE b.user_id = %s
    GROUP BY b.id, b.nombre_banco, b.monto
    ORDER BY b.id DESC
""", (session['user_id'],))
    bancos = cursor.fetchall()

    # Traer el gasto específico que se va a pagar, validando que pertenece al usuario logueado y que es un gasto (tipo_id=2).
    cursor.execute("""
        SELECT m.*, c.nombre AS categoria
        FROM movimientos m
        LEFT JOIN categorias c ON m.categoria_id = c.id
        WHERE m.id=%s AND m.user_id=%s AND m.tipo_id=2
    """, (id, session['user_id']))
    movimiento = cursor.fetchone()
    if not movimiento:
        flash("Movimiento no encontrado o no es un gasto", "danger")
        return redirect('/mov')

    # Determinar el banco seleccionado para el pago, ya sea por POST (cuando se envía el formulario) o por GET (cuando se carga la página por primera vez).
    banco_id = None
    if request.method == 'POST':
        banco_id = request.form.get('banco_id')
    elif bancos:
        banco_id = movimiento['banco_id'] if movimiento and movimiento['banco_id'] else bancos[0]['id']
    # Obtener el saldo real del banco seleccionado para mostrar en la página de pago.
    saldo_banco = None
    if banco_id:
        banco_match = next((b for b in bancos if str(b['id']) == str(banco_id)), None)
        saldo_banco = banco_match['saldo_real'] if banco_match else None
    # Validar que el banco seleccionado para el pago pertenece al usuario logueado y tiene saldo suficiente para cubrir el monto a pagar.
    if request.method == 'POST':
        try:
            monto_pagado = float(request.form.get('monto', '0').replace(',', '.'))
        except Exception:
            monto_pagado = 0

        # VALIDACIONES:
            # Validar que el banco seleccionado para el pago pertenece al usuario logueado.
        if not banco_match:
            flash("Banco no encontrado", "danger")
            return redirect(f'/mov/pago/{id}')
            # Validar que el banco seleccionado tiene saldo suficiente para cubrir el monto a pagar.
        if banco_match['saldo_real'] < monto_pagado:
            flash("Saldo insuficiente en el banco seleccionado", "danger")
            return redirect(f'/mov/pago/{id}')
            # Validar que el monto a pagar sea mayor a 0 y no supere el monto pendiente del gasto.
        if monto_pagado <= 0:
            flash("El monto a pagar debe ser mayor a 0", "danger")
            return redirect(f'/mov/pago/{id}')
            # Validar que el monto a pagar no supere el monto pendiente del gasto.
        if monto_pagado > movimiento['monto']:
            flash("No puedes pagar más que el monto del gasto", "danger")
            return redirect(f'/mov/pago/{id}')


        # Registrar el pago en historial_pagos
        cursor.execute("""
            INSERT INTO historial_pagos (
                    id_movimiento, 
                    fecha_pago, 
                    monto_pago, 
                    usuario)
            VALUES (%s, NOW(), %s, %s)
        """, (id, monto_pagado, session['user_id']))

        # Descontar el saldo del banco seleccionado
        cursor.execute("""
            UPDATE banco SET monto = monto - %s WHERE id = %s AND user_id = %s
        """, (monto_pagado, banco_id, session['user_id']))

        conn.commit()


        # Actualizar el monto pagado y el saldo pendiente del gasto
        cursor.execute("SELECT saldo_pendiente FROM movimientos WHERE id=%s AND user_id=%s", (id, session['user_id']))
        row = cursor.fetchone()
        # Si el saldo pendiente es 0 o menor, eliminar el gasto. Si no, mostrar mensaje de pago parcial registrado.
        if row and row['saldo_pendiente'] <= 0:
            cursor.execute("DELETE FROM historial_pagos WHERE id_movimiento=%s", (id,))
            cursor.execute("DELETE FROM movimientos WHERE id=%s AND user_id=%s", (id, session['user_id']))
            conn.commit()
            flash("Pago realizado y gasto eliminado correctamente", "success")
        # Si el saldo pendiente es mayor a 0, mostrar mensaje de pago parcial registrado.
        else:
            flash("Pago parcial registrado correctamente", "success")
        return redirect('/mov')

    # Traer nuevamente el listado de gastos del usuario para mostrar en la página de pago, ya que puede haber cambios en el monto pendiente después de realizar un pago.
    cursor.execute("""
        SELECT m.id, m.descripcion, m.monto, m.monto_pagado, m.saldo_pendiente
        FROM movimientos m
        WHERE m.user_id = %s AND m.tipo_id = 2
        ORDER BY m.fecha DESC, m.id DESC
    """, (session['user_id'],))
    gastos = cursor.fetchall()

    return render_template('pago.html', movimiento=movimiento, bancos=bancos, gastos=gastos, saldo_banco=saldo_banco)


# =========================
# MOVIMIENTOS DE TRANSFERENCIA
# =========================

@app.route('/mov/transferencia', methods=['POST'])
def crear_transferencia():
    if 'user_id' not in session:
        return redirect('/login')

    banco_origen = request.form['banco_origen']
    banco_destino = request.form['banco_destino']
    monto = float(request.form['monto_origen_transferencia'])
    fecha = request.form['fecha']

    conn = get_db_connection()
    cursor = conn.cursor()

    # 1. Validar que el banco origen tenga saldo suficiente para la transferencia.
    cursor.execute("""
        SELECT COALESCE(
            b.monto + SUM(
                CASE
                    WHEN m.tipo_id = 1 THEN m.monto
                    WHEN m.tipo_id = 2 THEN -m.monto
                    ELSE 0
                END
            ),
            b.monto
        ) AS saldo_actual
        FROM banco b
        LEFT JOIN movimientos m ON m.banco_id = b.id
        WHERE b.id = %s AND b.user_id = %s
        GROUP BY b.id, b.monto
    """, (banco_origen, session['user_id']))
    row = cursor.fetchone()
    saldo_disponible = row['saldo_actual'] if row else 0

    if monto > saldo_disponible:
        flash("El monto de la transferencia no puede ser mayor al saldo disponible del banco origen.", "danger")
        return redirect('/mov/transferencia')

    # 2. Insertar movimientos (origen negativo, destino positivo)
    cursor.execute("""
        INSERT INTO movimientos (user_id, monto, categoria_id, banco_id, tipo_id, fecha, descripcion)
        VALUES (%s, %s, NULL, %s, 3, %s, %s)
    """, (
        session['user_id'],
        -monto,
        banco_origen,
        fecha,
        f"Transferencia a banco {banco_destino}"
    ))
    # El movimiento de destino se registra como ingreso (monto positivo) para que sume al saldo del banco destino.
    cursor.execute("""
        INSERT INTO movimientos (user_id, monto, categoria_id, banco_id, tipo_id, fecha, descripcion)
        VALUES (%s, %s, NULL, %s, 3, %s, %s)
    """, (
        session['user_id'],
        monto,
        banco_destino,
        fecha,
        f"Transferencia desde banco {banco_origen}"
    ))

    # 3. Actualizar saldos directamente en tabla banco
    cursor.execute("UPDATE banco SET monto = monto - %s WHERE id=%s AND user_id=%s",
                   (monto, banco_origen, session['user_id']))
    cursor.execute("UPDATE banco SET monto = monto + %s WHERE id=%s AND user_id=%s",
                   (monto, banco_destino, session['user_id']))

    conn.commit()
    conn.close()

    flash("Transferencia realizada correctamente", "success")
    return redirect('/mov')






# =========================
# CATEGORIAS
# =========================
@app.route('/categorias')
def categorias():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM categorias")
    categorias = cursor.fetchall()
    return render_template('categorias.html', categorias=categorias)

@app.route('/categorias/crear', methods=['POST'])
def crear_categoria():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO categorias (nombre) VALUES (%s)", (request.form['nombre'],))
    conn.commit()
    return redirect('/categorias')

@app.route('/categorias/editar/<int:id>', methods=['GET', 'POST'])
def editar_categoria(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        cursor.execute(
            "UPDATE categorias SET nombre = %s WHERE id = %s",
            (request.form['nombre'], id)
        )
        conn.commit()
        return redirect('/categorias')

    cursor.execute("SELECT * FROM categorias WHERE id = %s", (id,))
    categoria = cursor.fetchone()

    return render_template('editar_categoria.html', categoria=categoria)

@app.route('/categorias/eliminar/<int:id>', methods=['POST'])
def eliminar_categoria(id):
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM movimientos WHERE categoria_id=%s",(id,))
    count = cursor.fetchone()[0]

    if count > 0:
        flash("No puedes eliminar una categoría en uso", "danger")
        return redirect('/categorias')

    cursor.execute("DELETE FROM categorias WHERE id=%s",(id,))
    conn.commit()

    return redirect('/categorias')

# =========================
# FAMILIA
# =========================
@app.route('/familia', methods=['GET', 'POST'])
def familia():
    if 'user_id' not in session:
        return redirect('/login')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        cursor.execute("INSERT INTO familia (nombre) VALUES (%s)", (request.form['nombre'],))
        conn.commit()

        familia_id = cursor.lastrowid

        cursor.execute(
            "UPDATE usuarios SET familia_id=%s WHERE id=%s",
            (familia_id, session['user_id'])
        )
        conn.commit()

        return redirect('/')

    cursor.execute("""
        SELECT f.* FROM familia f
        JOIN usuarios u ON u.familia_id=f.id
        WHERE u.id=%s
    """,(session['user_id'],))

    familia = cursor.fetchone()

    miembros = []
    if familia:
        cursor.execute("SELECT nombre,email FROM usuarios WHERE familia_id=%s",(familia['id'],))
        miembros = cursor.fetchall()

    return render_template('familia.html', familia=familia, miembros=miembros)

#Unirse a Familia
@app.route('/familia/unirse', methods=['POST'])
def unirse_familia():
    if 'user_id' not in session:
        return redirect('/login')

    familia_id = request.form.get('familia_id')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM familia WHERE id=%s", (familia_id,))
    familia = cursor.fetchone()

    if not familia:
        flash("La familia no existe", "danger")
        return redirect('/')

    cursor.execute("""
        UPDATE usuarios SET familia_id=%s WHERE id=%s
    """, (familia_id, session['user_id']))

    conn.commit()
    cursor.close()
    conn.close()

    flash("Te uniste a la familia", "success")
    return redirect('/')

#Salir de familia

@app.route('/familia/salir', methods=['POST'])
def salir_familia():
    if 'user_id' not in session:
        return redirect('/login')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE usuarios SET familia_id=NULL WHERE id=%s
    """, (session['user_id'],))

    conn.commit()
    cursor.close()
    conn.close()

    flash("Saliste de la familia", "success")
    return redirect('/familia')


# =========================
# BANCOS
# =========================

# 🔥 AHORA SOPORTA /banco Y /banco/<id>
@app.route('/banco', methods=['GET'])
@app.route('/banco/<int:id>', methods=['GET', 'POST'])
def banco(id=None):

    if 'user_id' not in session:
        return redirect('/login')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # =========================
    # EDITAR BANCO (POST)
    # =========================
    if id and request.method == 'POST':

        nombre_banco = request.form.get('nombre_banco', '').strip()

        monto_str = request.form.get('saldo_inicial', '0')
        monto_str = monto_str.replace('.', '').replace(',', '')
        monto = float(monto_str) if monto_str else 0

        cursor.execute("""
            UPDATE banco
            SET nombre_banco = %s, monto = %s
            WHERE id = %s AND user_id = %s
        """, (nombre_banco, monto, id, session['user_id']))

        conn.commit()
        cursor.close()
        conn.close()

        flash("Banco actualizado correctamente", "success")
        return redirect('/banco')

    # =========================
    # MOSTRAR EDITAR
    # =========================
    if id:
        cursor.execute("""
            SELECT * FROM banco
            WHERE id = %s AND user_id = %s
        """, (id, session['user_id']))

        banco = cursor.fetchone()

        cursor.close()
        conn.close()

        if not banco:
            flash("Banco no encontrado", "danger")
            return redirect('/banco')

        return render_template('editar_banco.html', banco=banco)

    # =========================
    # LISTADO (usa tu función actual)
    # =========================
    cursor.close()
    conn.close()
    return bancos()


# =========================
# LISTADO BANCOS
# =========================
@app.route('/bancos', methods=['GET'])
def bancos():
    if 'user_id' not in session:
        return redirect('/login')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
                
        cursor.execute("""
    SELECT
        b.id,
        b.nombre_banco AS nombre,
        b.monto,
        COALESCE(
            b.monto + SUM(
                CASE
                    WHEN m.tipo_id = 1 THEN m.monto
                    WHEN m.tipo_id = 2 THEN -m.monto
                    ELSE 0
                END
            ),
            b.monto
        ) AS saldo_actual
    FROM banco b
    LEFT JOIN movimientos m ON m.banco_id = b.id
    WHERE b.user_id = %s
    GROUP BY b.id, b.nombre_banco, b.monto
    ORDER BY b.id DESC
""", (session['user_id'],))
        
        bancos = cursor.fetchall()

    except mysql.connector.Error:
        bancos = []

    # 🔥 TIPOS
    cursor.execute("SELECT id, nombre FROM tipo_banco ORDER BY nombre ASC")
    tipos_banco = cursor.fetchall()

    cursor.execute("SELECT id, nombre FROM tipo_cuenta ORDER BY nombre ASC")
    tipos_cuenta = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        'banco.html',
        bancos=bancos,
        tipos_banco=tipos_banco,
        tipos_cuenta=tipos_cuenta
    )


# =========================
# CREAR BANCO
# =========================
@app.route('/banco/crear', methods=['POST'])
def crear_banco():
    if 'user_id' not in session:
        return redirect('/login')

    conn = get_db_connection()
    cursor = conn.cursor()

    tipo_banco_id = request.form.get('tipo_banco_id')
    tipo_cuenta_id = request.form.get('tipo_cuenta_id')

    nombre_banco = request.form.get('nombre_banco', '').strip()

    monto_str = request.form.get('monto', '0')
    monto_str = monto_str.replace('.', '').replace(',', '')
    saldo_inicial = float(monto_str) if monto_str else 0

    if not tipo_banco_id:
        cursor.execute("SELECT id FROM tipo_banco LIMIT 1")
        row = cursor.fetchone()
        tipo_banco_id = row[0] if row else None

    if not tipo_cuenta_id:
        cursor.execute("SELECT id FROM tipo_cuenta LIMIT 1")
        row = cursor.fetchone()
        tipo_cuenta_id = row[0] if row else None

    if not tipo_banco_id or not tipo_cuenta_id:
        flash("Debes crear tipos de banco y cuenta primero", "danger")
        return redirect('/banco')

    cursor.execute("""
        INSERT INTO banco (user_id, tipo_banco_id, tipo_cuenta_id, nombre_banco, monto)
        VALUES (%s, %s, %s, %s, %s)
    """, (
        session['user_id'],
        tipo_banco_id,
        tipo_cuenta_id,
        nombre_banco,
        saldo_inicial
    ))

    conn.commit()
    cursor.close()
    conn.close()

    flash("Banco creado correctamente", "success")
    return redirect('/')


# =========================
# ELIMINAR BANCO
# =========================
@app.route('/banco/eliminar', methods=['POST'])
def eliminar_banco():
    if 'user_id' not in session:
        return redirect('/login')

    banco_id = request.form.get('id')

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("DELETE FROM banco WHERE id=%s AND user_id=%s",
                       (banco_id, session['user_id']))
        conn.commit()
        flash("Banco eliminado correctamente", "success")
    except mysql.connector.Error:
        flash("No se puede eliminar (tiene movimientos)", "danger")

    cursor.close()
    conn.close()

    return redirect('/banco')


# =========================
# TIPOS DE BANCO
# =========================
@app.route('/tipo_banco', methods=['GET'])
def tipo_banco():
    if 'user_id' not in session:
        return redirect('/login')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT id, nombre FROM tipo_banco")
    tipos_banco = cursor.fetchall()

    return render_template('tipo_banco.html', tipos_banco=tipos_banco)

# =========================
# PERFIL
# =========================
@app.route('/perfil', methods=['GET', 'POST'])
def perfil():
    if 'user_id' not in session:
        return redirect('/login')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        cursor.execute("""
            UPDATE usuarios
            SET nombre=%s, email=%s, password=%s
            WHERE id=%s
        """,(request.form['nombre'],request.form['email'],request.form['password'],session['user_id']))

        conn.commit()
        session['usuario_nombre'] = request.form['nombre']
        flash("Perfil actualizado correctamente", "success")

        return redirect('/')

    cursor.execute("SELECT * FROM usuarios WHERE id=%s",(session['user_id'],))
    usuario = cursor.fetchone()

    return render_template('perfil.html', usuario=usuario)

# =========================
# REPORTES
# =========================

#Endpoint genericos para reportes.

# vista principal de reportes, con links a cada reporte específico.
@app.route("/reportes")
def reportes():
    if "user_id" not in session:
        return redirect("/login")
    return render_template("reportes.html")

# Visualizar reportes.
@app.route("/reportes/visualizar")
def visualizar_reportes():
    if "user_id" not in session:
        return redirect("/login")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    conn.close()

    return render_template("visualizar_reportes.html", reportes=reportes)

#Visualizar reporte en PDF.
@app.route("/reportes/visualizar/pdf")
def visualizar_reportes_pdf():
    if "user_id" not in session:
        return redirect("/login")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    conn.close()

    return render_template("visualizar_reportes_pdf.html", reportes=reportes)

#--------------------UTILIDAD-----------------------
def obtener_filas_reportes_utilidad(user_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT 
            DATE_FORMAT(fecha, '%m-%Y') AS Mes,
            SUM(CASE WHEN tipo_id = 1 THEN monto ELSE 0 END) AS Ingresos,
            SUM(CASE WHEN tipo_id = 2 THEN monto ELSE 0 END) AS Gastos,
            (SUM(CASE WHEN tipo_id = 1 THEN monto ELSE 0 END) -
             SUM(CASE WHEN tipo_id = 2 THEN monto ELSE 0 END)) AS Utilidad
        FROM movimientos
        WHERE user_id = %s
        GROUP BY Mes
        ORDER BY STR_TO_DATE(Mes, '%m-%Y') DESC
    """, (user_id,))
    filas = cursor.fetchall()
    conn.close()
    return filas

# EXCEL

@app.route("/reporte/utilidad/excel")
def reporte_utilidad_excel():
    if "user_id" not in session:
        return redirect("/login")

    filas = obtener_filas_reportes_utilidad(session["user_id"])
    # Crear un libro de Excel y una hoja.
    wb = Workbook()
    ws = wb.active
    ws.title = "Utilidad Mensual"

    # Primera fila vacía para que el encabezado inicie en B2.
    ws.append(["","", "Reporte de Utilidad Mensual"])
    ws.cell(row=ws.max_row, column=3).font = Font(size=16, bold=True)
    ws.append([])
    ws.append(["","Usuario:", session.get("usuario_nombre", "Desconocido"),"","Generado el:", datetime.now().strftime("%d-%m-%Y %H:%M")]) # Fecha de generación del reporte
    ws.append(["", "N°", "Mes", "Gastos", "Ingresos", "Utilidad"])
    header_row = ws.max_row

    resumen_por_mes = {f["Mes"]: float(f.get("Utilidad") or 0) for f in filas if f.get("Mes")}

    # Datos desde B3.
    for i, f in enumerate(filas, start=1):
        mes = f.get("Mes")
        gastos = float(f.get("Gastos") or 0)
        ingresos = float(f.get("Ingresos") or 0)
        utilidad = float(f.get("Utilidad") or 0)
        ws.append([
            "",
            i,
            mes,
            gastos,
            ingresos,
            utilidad
        ])
        ws.cell(row=ws.max_row, column=4).number_format = '#,##0'
        ws.cell(row=ws.max_row, column=5).number_format = '#,##0'
        ws.cell(row=ws.max_row, column=6).number_format = '#,##0'

    # Convertir en tabla (encabezado y datos).
    end_row = ws.max_row
    tabla = Table(displayName="TablaUtilidad", ref=f"B{header_row}:F{end_row}")
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight8", # Estilo de tabla.
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws.add_table(tabla)

    # Resumen al final del reporte: total de gastos por mes.
    ws.append([])
    ws.append([])
    ws.append(["Resumen del mes en curso:"])
    for celda in ws[ws.max_row]:
        celda.font = Font(size=14, bold=True)
    ws.append(["Mes", "Total Utilidad"])
    fila = ws.max_row
    ws.cell(row=fila, column=1).font = Font(bold=True)
    ws.cell(row=fila, column=2).font = Font(bold=True)
    mes_actual = datetime.now().strftime("%m-%Y")
    total_mes_actual = resumen_por_mes.get(mes_actual, 0)
    ws.append([mes_actual, total_mes_actual])
    ws.cell(row=ws.max_row, column=2).number_format = '#,##0'

    #Guardar el libro de Excel en BytesIO.
    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)

    # Descargar el archivo excel.
    return send_file(
        archivo,
        as_attachment=True,
        download_name="Utilidad.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# PDF

@app.route("/reporte/utilidad/pdf")
def reporte_utilidad_pdf():
    if "user_id" not in session:
        return redirect("/login")

    filas = obtener_filas_reportes_utilidad(session["user_id"])
    archivo = BytesIO()

    # Crear el documento PDF con ReportLab.
    doc = SimpleDocTemplate(
        archivo,
        pagesize=A4,
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    # Crear estilos y elementos para el PDF.
    styles = getSampleStyleSheet()
    elementos = [
        Paragraph("Reporte de Utilidad Mensual", styles["Title"]),
        Paragraph(f"Usuario: {session.get('usuario_nombre', 'Desconocido')}", styles["Normal"]),
        Paragraph(f"Generado el: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles["Normal"]),
        Spacer(1, 12),
    ]
    # Titulos de la tabla.
    datos_tabla = [["N°", "Mes", "Gastos", "Ingresos", "Utilidad"]]
    utilidad_mes_actual = 0.0
    mes_actual = datetime.now().strftime("%m-%Y")
    # Datos de la tabla.
    for i, fila in enumerate(filas, start=1):
        mes = fila.get("Mes") or "-"
        gastos = float(fila.get("Gastos") or 0)
        ingresos = float(fila.get("Ingresos") or 0)
        utilidad = float(fila.get("Utilidad") or 0)

        if mes == mes_actual:
            utilidad_mes_actual = utilidad
        # Agregar fila a los datos de la tabla.
        datos_tabla.append([
            str(i),
            mes,
            formatear_miles(gastos),
            formatear_miles(ingresos),
            formatear_miles(utilidad),
        ])
    # Crear la tabla PDF con los datos.
    tabla = PdfTable(datos_tabla, repeatRows=1)
    tabla.setStyle(PdfTableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    # Aplicar diseño a la tabla.
    elementos.append(tabla)
    elementos.append(Spacer(1, 12))
    elementos.append(Paragraph("Resumen del mes en curso", styles["Heading3"]))
    elementos.append(Paragraph(f"Mes: {mes_actual}", styles["Normal"]))
    elementos.append(Paragraph(f"Total utilidad: {formatear_miles(utilidad_mes_actual)}", styles["Normal"]))

    doc.build(elementos)
    archivo.seek(0)

    return send_file(
        archivo,
        as_attachment=True,
        download_name="Utilidad.pdf",
        mimetype="application/pdf",
    )

#------------REPORTE GASTOS MENSUALES------------------

#Funcion que muentran los gastos menusuales del usuario, con su categoria y familia.
def obtener_filas_reporte_gastos_mensuales(user_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT
            c.nombre AS Categoria,
            m.descripcion AS Nombre,
            f.nombre AS Familia,
            DATE_FORMAT(m.fecha, '%m-%Y') AS Mes,
            m.monto AS Monto
        FROM movimientos m
        LEFT JOIN categorias c ON m.categoria_id = c.id
        LEFT JOIN usuarios u ON m.user_id = u.id
        LEFT JOIN familia f ON u.familia_id = f.id
        WHERE m.tipo_id = 2 AND m.user_id = %s
        ORDER BY Mes DESC
    """, (user_id,))
    filas = cursor.fetchall()
    conn.close()
    return filas

# Construir un resumen de gastos por mes a partir de las filas obtenidas.
def construir_resumen_por_mes(filas):
    resumen = {}
    for fila in filas:
        monto = float(fila["Monto"] or 0)
        mes = fila.get("Mes")
        if not mes:
            continue
        resumen[mes] = resumen.get(mes, 0) + monto

    def parsear_mes(valor):
        try:
            return datetime.strptime(valor, "%m-%Y")
        except (TypeError, ValueError):
            return datetime.min

    # Ordenar el resumen por mes (formato MM-YYYY) de forma descendente.
    return sorted(
        resumen.items(),
        key=lambda item: parsear_mes(item[0]),
        reverse=True
    )


# EXCEL

@app.route("/reporte/gastos_mensuales/excel")
def reporte_gastos_mensuales_excel():
    if "user_id" not in session:
        return redirect("/login")

    filas = obtener_filas_reporte_gastos_mensuales(session["user_id"])
    # Crear un libro de Excel y una hoja.
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos Mensuales"

    # Primera fila vacía para que el encabezado inicie en B2.
    ws.append(["","", "Reporte de Gastos Mensuales"])
    ws.cell(row=ws.max_row, column=3).font = Font(size=16, bold=True)
    ws.append([])
    ws.append(["","Usuario:", session.get("usuario_nombre", "Desconocido"),"","Generado el:", datetime.now().strftime("%d-%m-%Y %H:%M")]) # Fecha de generación del reporte
    ws.append(["", "N°", "Mes", "Nombre", "Categoria", "Familia", "Monto"])
    header_row = ws.max_row

    resumen_por_mes = dict(construir_resumen_por_mes(filas))

    # Datos desde B3.
    for i, f in enumerate(filas, start=1):
        monto = float(f["Monto"] or 0)
        mes = f["Mes"]
        ws.append([
            "",
            i,
            mes,
            f["Nombre"],
            f["Categoria"] if f["Categoria"] else "Sin categoria",
            f["Familia"] if f["Familia"] else "Sin familia",
            monto
        ])
        ws.cell(row=ws.max_row, column=7).number_format = '#,##0'
        resumen_por_mes[mes] = resumen_por_mes.get(mes, 0) + monto

    # Convertir datos en tabla (Titulos + datos).
    end_row = ws.max_row
    tabla = Table(displayName="TablaUtilidad", ref=f"B{header_row}:G{end_row}")
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight8", # Estilo de tabla.
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws.add_table(tabla)

    # Resumen al final del reporte: total de gastos por mes.
    ws.append([])
    ws.append([])
    ws.append(["Resumen por mes:"])
    for celda in ws[ws.max_row]:
        celda.font = Font(size=14, bold=True)
    ws.append(["Mes", "Total Gastos"])
    fila = ws.max_row
    ws.cell(row=fila, column=1).font = Font(bold=True)
    ws.cell(row=fila, column=2).font = Font(bold=True)
    for mes in sorted((m for m in resumen_por_mes.keys() if m), reverse=True):
        ws.append([mes, resumen_por_mes[mes]])
        ws.cell(row=ws.max_row, column=2).number_format = '#,##0'

    #Guardar el libro de Excel en BytesIO.
    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)

    # Descargar el archivo excel.
    return send_file(
        archivo,
        as_attachment=True,
        download_name="Gastos_mensuales.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# PDF


@app.route("/reporte/gastos_mensuales/pdf")
def reporte_gastos_mensuales_pdf():
    if "user_id" not in session:
        return redirect("/login")

    filas = obtener_filas_reporte_gastos_mensuales(session["user_id"])
    archivo = BytesIO()

    # Crear el documento PDF con ReportLab.
    doc = SimpleDocTemplate(
        archivo,
        pagesize=A4,
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    # Crear estilos y elementos para el PDF.
    styles = getSampleStyleSheet()
    elementos = [
        Paragraph("Reporte de Gastos Mensuales", styles["Title"]),
        Paragraph(f"Usuario: {session.get('usuario_nombre', 'Desconocido')}", styles["Normal"]),
        Paragraph(f"Generado el: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles["Normal"]),
        Spacer(1, 12),
    ]
    # Titulos de la tabla.
    datos_tabla = [["N°", "Mes", "Nombre", "Categoria", "Familia", "Monto"]]
    resumen_por_mes = {}
    # Datos de la tabla.
    for i, fila in enumerate(filas, start=1):
        mes = fila.get("Mes") or "-"
        nombre = str(fila.get("Nombre") or "-")
        categoria = str(fila.get("Categoria") or "")
        familia = str(fila.get("Familia") or "")
        monto = float(fila.get("Monto") or 0)

        if mes != "-":
            resumen_por_mes[mes] = resumen_por_mes.get(mes, 0) + monto

        # Agregar fila a los datos de la tabla.
        datos_tabla.append([
            str(i),
            mes,
            nombre,
            categoria,
            familia,
            formatear_miles(monto),
        ])
    # Crear la tabla PDF con los datos.
    tabla = PdfTable(datos_tabla, repeatRows=1)
    tabla.setStyle(PdfTableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    # Aplicar diseño a la tabla.
    elementos.append(tabla)
    elementos.append(Spacer(1, 12))
    elementos.append(Paragraph("Resumen por mes", styles["Heading3"]))
    # Construir una tabla con el resumen de gastos por mes.
    resumen_tabla = [["Mes", "Total Gastos"]]
    for mes in sorted((m for m in resumen_por_mes.keys() if m), reverse=True):
        resumen_tabla.append([mes, formatear_miles(resumen_por_mes[mes])])

    tabla_resumen = PdfTable(resumen_tabla, repeatRows=1)
    tabla_resumen.setStyle(PdfTableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elementos.append(tabla_resumen)

    doc.build(elementos)
    archivo.seek(0)

    return send_file(
        archivo,
        as_attachment=True,
        download_name="Gastos.pdf",
        mimetype="application/pdf",
    )



#------------REPORTE INGRESOS MENSUALES------------------

#Funcion que muentran los gastos menusuales del usuario, con su categoria y familia.
def obtener_filas_reporte_ingresos_mensuales(user_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT
            c.nombre AS Categoria,
            m.descripcion AS Nombre,
            f.nombre AS Familia,
            DATE_FORMAT(m.fecha, '%m-%Y') AS Mes,
            m.monto AS Monto
        FROM movimientos m
        LEFT JOIN categorias c ON m.categoria_id = c.id
        LEFT JOIN usuarios u ON m.user_id = u.id
        LEFT JOIN familia f ON u.familia_id = f.id
        WHERE m.tipo_id = 1 AND m.user_id = %s
        ORDER BY Mes DESC
    """, (user_id,))
    filas = cursor.fetchall()
    conn.close()
    return filas


# EXCEL

@app.route("/reporte/ingresos_mensuales/excel")
def reporte_ingresos_mensuales_excel():
    if "user_id" not in session:
        return redirect("/login")

    filas = obtener_filas_reporte_ingresos_mensuales(session["user_id"])
    # Crear un libro de Excel y una hoja.
    wb = Workbook()
    ws = wb.active
    ws.title = "Ingresos Mensuales"

    # Primera fila vacía para que el encabezado inicie en B2.
    ws.append(["","", "Reporte de Ingresos Mensuales"])
    ws.cell(row=ws.max_row, column=3).font = Font(size=16, bold=True)
    ws.append([])
    ws.append(["","Usuario:", session.get("usuario_nombre", "Desconocido"),"","Generado el:", datetime.now().strftime("%d-%m-%Y %H:%M")]) # Fecha de generación del reporte
    ws.append(["", "N°", "Mes", "Nombre", "Categoria", "Familia", "Monto"])
    header_row = ws.max_row

    resumen_por_mes = dict(construir_resumen_por_mes(filas))

    # Datos desde B3.
    for i, f in enumerate(filas, start=1):
        monto = float(f["Monto"] or 0)
        mes = f["Mes"]
        ws.append([
            "",
            i,
            mes,
            f["Nombre"],
            f["Categoria"] if f["Categoria"] else "Sin categoria",
            f["Familia"] if f["Familia"] else "Sin familia",
            monto
        ])
        ws.cell(row=ws.max_row, column=7).number_format = '#,##0'
        resumen_por_mes[mes] = resumen_por_mes.get(mes, 0) + monto

    # Convertir bloque principal en tabla (encabezado + datos).
    end_row = ws.max_row
    tabla = Table(displayName="TablaIngresos", ref=f"B{header_row}:G{end_row}")
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight8", # Estilo de tabla.
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws.add_table(tabla)

    # Resumen al final del reporte: total de gastos por mes.
    ws.append([])
    ws.append([])
    ws.append(["Resumen por mes:"])
    for celda in ws[ws.max_row]:
        celda.font = Font(size=14, bold=True)
    ws.append(["Mes", "Total Ingresos"])
    fila = ws.max_row
    ws.cell(row=fila, column=1).font = Font(bold=True)
    ws.cell(row=fila, column=2).font = Font(bold=True)
    for mes in sorted((m for m in resumen_por_mes.keys() if m), reverse=True):
        ws.append([mes, resumen_por_mes[mes]])
        ws.cell(row=ws.max_row, column=2).number_format = '#,##0'

    #Guardar el libro de Excel en BytesIO.
    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)

    # Descargar el archivo excel.
    return send_file(
        archivo,
        as_attachment=True,
        download_name="Ingresos_mensuales.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# PDF

@app.route("/reporte/ingresos_mensuales/pdf")
def reporte_ingresos_mensuales_pdf():
    if "user_id" not in session:
        return redirect("/login")

    filas = obtener_filas_reporte_ingresos_mensuales(session["user_id"])
    archivo = BytesIO()

    # Crear el documento PDF con ReportLab.
    doc = SimpleDocTemplate(
        archivo,
        pagesize=A4,
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    # Crear estilos y elementos para el PDF.
    styles = getSampleStyleSheet()
    elementos = [
        Paragraph("Reporte de Ingresos Mensuales", styles["Title"]),
        Paragraph(f"Usuario: {session.get('usuario_nombre', 'Desconocido')}", styles["Normal"]),
        Paragraph(f"Generado el: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles["Normal"]),
        Spacer(1, 12),
    ]
    # Titulos de la tabla.
    datos_tabla = [["N°", "Mes", "Nombre", "Categoria", "Familia", "Monto"]]
    resumen_por_mes = {}
    # Datos de la tabla.
    for i, fila in enumerate(filas, start=1):
        mes = fila.get("Mes") or "-"
        nombre = str(fila.get("Nombre") or "-")
        categoria = str(fila.get("Categoria") or "")
        familia = str(fila.get("Familia") or "")
        monto = float(fila.get("Monto") or 0)

        if mes != "-":
            resumen_por_mes[mes] = resumen_por_mes.get(mes, 0) + monto

        # Agregar fila a los datos de la tabla.
        datos_tabla.append([
            str(i),
            mes,
            nombre,
            categoria,
            familia,
            formatear_miles(monto),
        ])
    # Crear la tabla PDF con los datos.
    tabla = PdfTable(datos_tabla, repeatRows=1)
    tabla.setStyle(PdfTableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    # Aplicar diseño a la tabla.
    elementos.append(tabla)
    elementos.append(Spacer(1, 12))
    elementos.append(Paragraph("Resumen por mes", styles["Heading3"]))
    # Construir una tabla con el resumen de ingresos por mes.
    resumen_tabla = [["Mes", "Total Ingresos"]]
    for mes in sorted((m for m in resumen_por_mes.keys() if m), reverse=True):
        resumen_tabla.append([mes, formatear_miles(resumen_por_mes[mes])])

    tabla_resumen = PdfTable(resumen_tabla, repeatRows=1)
    tabla_resumen.setStyle(PdfTableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elementos.append(tabla_resumen)

    doc.build(elementos)
    archivo.seek(0)

    return send_file(
        archivo,
        as_attachment=True,
        download_name="Ingresos_mensuales.pdf",
        mimetype="application/pdf",
    )


@app.route('/reset/<token>', methods=['GET', 'POST'])
def reset_password(token):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT * FROM usuarios 
        WHERE reset_token=%s AND reset_expira > NOW()
    """, (token,))
    user = cursor.fetchone()

    if not user:
        flash("Token inválido o expirado", "danger")
        return redirect('/login')

    if request.method == 'POST':
        nueva_password = request.form['password']

        cursor.execute("""
            UPDATE usuarios 
            SET password=%s, reset_token=NULL, reset_expira=NULL
            WHERE id=%s
        """, (nueva_password, user['id']))

        conn.commit()
        cursor.close()
        conn.close()

        flash("Contraseña actualizada correctamente", "success")
        return redirect('/login')

    return render_template('reset.html')



# =========================
# RUN
# =========================
if __name__ == '__main__':
    app.run(debug=True)
